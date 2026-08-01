import streamlit as st
import pandas as pd
from datetime import date, time, timedelta, datetime
from database import get_connection, execute, read_sql_query, get_config, get_ubicaciones

st.set_page_config(page_title="Registro de Horario Semanal", page_icon="🕒", layout="centered")

nombre_empresa = get_config("nombre_empresa", "Mi Empresa")
st.title(f"🕒 Registro de Horario Semanal")
st.caption(f"{nombre_empresa} — solo entrada y salida, se guarda directo en el ERP")

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

# ---------- LOGIN BÁSICO (opcional, mismo usuario del ERP si se configura) ----------
def check_login():
    if "usuarios" not in st.secrets:
        return True  # si no se configuró login, la app queda abierta (útil para tablet compartida)
    if "autenticado" not in st.session_state:
        st.session_state.autenticado = False
    if st.session_state.autenticado:
        return True

    with st.form("login_form"):
        usuario = st.text_input("Usuario")
        clave = st.text_input("Contraseña", type="password")
        enviar = st.form_submit_button("Ingresar")
    if enviar:
        usuarios_validos = st.secrets.get("usuarios", {})
        if usuario in usuarios_validos and clave == usuarios_validos[usuario]:
            st.session_state.autenticado = True
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos")
    return False


if not check_login():
    st.stop()

try:
    horas_normales_max = float(get_config("horas_normales_por_dia", "7"))
except Exception as e:
    st.error(f"No se pudo conectar a la base de datos: {e}")
    st.stop()

# ---------- CARGAR EMPLEADOS ----------
with get_connection() as conn:
    empleados_df = read_sql_query(
        "SELECT * FROM empleados WHERE estado='Activo' ORDER BY nombre", conn
    )

# ---------- AGREGAR EMPLEADO NUEVO ----------
with st.expander("➕ Agregar un empleado nuevo"):
    ubicaciones = get_ubicaciones()
    with st.form("nuevo_empleado_form", clear_on_submit=True):
        nombre_nuevo = st.text_input("Nombre completo *")
        if ubicaciones:
            ubicacion_id_nuevo = st.selectbox(
                "Sede", [u["id"] for u in ubicaciones],
                format_func=lambda x: next(u["nombre"] for u in ubicaciones if u["id"] == x)
            )
        else:
            st.warning("No hay sedes creadas todavía en el ERP — el empleado se creará sin sede asignada.")
            ubicacion_id_nuevo = None
        crear = st.form_submit_button("Agregar empleado")

    if crear:
        if not nombre_nuevo.strip():
            st.error("El nombre es obligatorio.")
        else:
            with get_connection() as conn:
                execute(conn, """
                    INSERT INTO empleados (nombre, ubicacion_id, estado, salario_base, valor_hora)
                    VALUES (?, ?, 'Activo', 0, 0)
                """, (nombre_nuevo.strip(), ubicacion_id_nuevo))
            st.success(f"Empleado '{nombre_nuevo}' agregado. Complétalo con más datos (salario, cédula, afiliaciones) directamente en el ERP cuando puedas.")
            st.rerun()

if empleados_df.empty:
    st.warning("Todavía no hay empleados activos. Agrega el primero arriba.")
    st.stop()

st.divider()

# ---------- SELECCIÓN DE EMPLEADO Y SEMANA ----------
empleado_id = st.selectbox(
    "Empleado",
    empleados_df["id"].tolist(),
    format_func=lambda x: empleados_df[empleados_df["id"] == x]["nombre"].values[0]
)

hoy = date.today()
lunes_actual = hoy - timedelta(days=hoy.weekday())
fecha_inicio_semana = st.date_input("Lunes de la semana a registrar", value=lunes_actual)

# ---------- CARGAR HORAS YA GUARDADAS DE ESA SEMANA (para no perderlas al recargar) ----------
fecha_fin_semana = fecha_inicio_semana + timedelta(days=6)
with get_connection() as conn:
    horas_existentes = read_sql_query("""
        SELECT * FROM horas WHERE empleado_id=? AND fecha BETWEEN ? AND ?
    """, conn, params=(int(empleado_id), str(fecha_inicio_semana), str(fecha_fin_semana)))
existentes_por_fecha = {row["fecha"]: row for _, row in horas_existentes.iterrows()} if not horas_existentes.empty else {}

st.subheader("Horario de la semana")
st.caption("Marca 'No trabajó' en los días libres. El resto se guarda con la hora de entrada y salida que pongas.")

valores_dia = {}
for i, dia_nombre in enumerate(DIAS):
    fecha_dia = fecha_inicio_semana + timedelta(days=i)
    fecha_str = str(fecha_dia)
    existente = existentes_por_fecha.get(fecha_str)

    col_dia, col_libre, col_entrada, col_salida = st.columns([2, 1.3, 1.3, 1.3])
    with col_dia:
        st.markdown(f"**{dia_nombre}**")
        st.caption(fecha_dia.strftime("%d/%m/%Y"))

    entrada_default = time(8, 0)
    salida_default = time(17, 0)
    if existente is not None:
        try:
            entrada_default = datetime.strptime(str(existente["hora_entrada"])[:5], "%H:%M").time()
            salida_default = datetime.strptime(str(existente["hora_salida"])[:5], "%H:%M").time()
        except Exception:
            pass
        # Ya hay un registro guardado: se respeta si tiene horas reales o no
        libre_default = float(existente["horas_normales"] or 0) == 0 and not existente["hora_entrada"]
    else:
        # Sin registro todavía: el domingo viene marcado como no laboral por defecto
        libre_default = (dia_nombre == "Domingo")

    with col_libre:
        no_trabajo = st.checkbox("No trabajó", value=libre_default, key=f"libre_{i}")
    with col_entrada:
        h_entrada = st.time_input("Entrada", value=entrada_default, key=f"entrada_{i}", disabled=no_trabajo, label_visibility="collapsed")
    with col_salida:
        h_salida = st.time_input("Salida", value=salida_default, key=f"salida_{i}", disabled=no_trabajo, label_visibility="collapsed")

    valores_dia[fecha_str] = {"no_trabajo": no_trabajo, "entrada": h_entrada, "salida": h_salida}
    st.divider()

if st.button("💾 Guardar horario de la semana", type="primary", use_container_width=True):
    guardados = 0
    with get_connection() as conn:
        for fecha_str, datos in valores_dia.items():
            if datos["no_trabajo"]:
                continue

            entrada_dt = datetime.combine(date.today(), datos["entrada"])
            salida_dt = datetime.combine(date.today(), datos["salida"])
            tiempo_bruto = max(0, (salida_dt - entrada_dt).total_seconds() / 3600)
            horas_normales = min(tiempo_bruto, horas_normales_max)

            existente = existentes_por_fecha.get(fecha_str)
            if existente is not None:
                execute(conn, """
                    UPDATE horas SET hora_entrada=?, hora_salida=?, horas_normales=?
                    WHERE id=?
                """, (str(datos["entrada"]), str(datos["salida"]), round(horas_normales, 4), int(existente["id"])))
            else:
                execute(conn, """
                    INSERT INTO horas (empleado_id, fecha, hora_entrada, hora_salida, horas_normales,
                    horas_extra_diurna, horas_extra_nocturna, horas_extra_dominical_festivo,
                    horas_extra_dominical_festivo_nocturna, horas_recargo_nocturno, horas_recargo_dominical,
                    horas_recargo_dominical_festivo_nocturno, horas_descuento, bonificacion, deduccion)
                    VALUES (?, ?, ?, ?, ?, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
                """, (int(empleado_id), fecha_str, str(datos["entrada"]), str(datos["salida"]), round(horas_normales, 4)))
            guardados += 1

    st.success(f"Horario guardado: {guardados} día(s) registrados para la semana del {fecha_inicio_semana.strftime('%d/%m/%Y')}.")
    st.info("Las horas extra, recargos, tiempo a descontar y demás detalles de nómina se completan en el ERP principal.")

# =========================================================
# SEMANA COMPLETA — TODOS LOS EMPLEADOS: editar, eliminar y descargar
# =========================================================
st.divider()
st.subheader("📋 Semana completa (todos los empleados)")
st.caption(f"Del {fecha_inicio_semana.strftime('%d/%m/%Y')} al {fecha_fin_semana.strftime('%d/%m/%Y')}")

with get_connection() as conn:
    semana_df = read_sql_query("""
        SELECT h.id, h.empleado_id, e.nombre AS empleado, h.fecha, h.hora_entrada, h.hora_salida, h.horas_normales
        FROM horas h JOIN empleados e ON h.empleado_id = e.id
        WHERE h.fecha BETWEEN ? AND ?
        ORDER BY h.fecha, e.nombre
    """, conn, params=(str(fecha_inicio_semana), str(fecha_fin_semana)))

if semana_df.empty:
    st.info("Todavía no hay ningún registro guardado para esta semana.")
else:
    dias_semana_map = {i: nombre for i, nombre in enumerate(DIAS)}

    def nombre_dia(fecha_str):
        fecha_obj = datetime.strptime(str(fecha_str)[:10], "%Y-%m-%d").date()
        return dias_semana_map[fecha_obj.weekday()]

    semana_df["dia"] = semana_df["fecha"].apply(nombre_dia)

    def calcular_horas_brutas(row):
        try:
            h_ent = datetime.strptime(str(row["hora_entrada"])[:5], "%H:%M")
            h_sal = datetime.strptime(str(row["hora_salida"])[:5], "%H:%M")
            return round(max(0, (h_sal - h_ent).total_seconds() / 3600), 4)
        except Exception:
            return 0.0

    semana_df["horas_brutas"] = semana_df.apply(calcular_horas_brutas, axis=1)

    tabla_mostrar = semana_df[["empleado", "dia", "fecha", "hora_entrada", "hora_salida", "horas_brutas", "horas_normales"]]
    tabla_mostrar.columns = ["Empleado", "Día", "Fecha", "Entrada", "Salida", "Horas brutas", "Horas normales"]
    st.dataframe(tabla_mostrar, use_container_width=True, hide_index=True)

    st.markdown("**Total de horas brutas de la semana por empleado**")
    resumen_brutas = semana_df.groupby("empleado")["horas_brutas"].sum().reset_index()
    resumen_brutas.columns = ["Empleado", "Total horas brutas semana"]
    st.dataframe(resumen_brutas, use_container_width=True, hide_index=True)

    # ---------- Descargar en Excel: hoja Resumen + una hoja por empleado ----------
    import io
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    azul_oscuro = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    amarillo = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fuente_header = Font(bold=True, color="FFFFFF")
    fuente_negrita = Font(bold=True)
    centrado = Alignment(horizontal="center", vertical="center")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Hoja Resumen: total de horas brutas y normales por empleado
        resumen_completo = semana_df.groupby("empleado").agg(
            horas_brutas=("horas_brutas", "sum"),
            horas_normales=("horas_normales", "sum")
        ).reset_index()
        resumen_completo.columns = ["Empleado", "Total horas brutas", "Total horas normales"]
        resumen_completo.to_excel(writer, sheet_name="Resumen", index=False)

        # Una hoja por empleado, con sus días y el total al final
        nombres_hoja_usados = set()
        for nombre_emp, grupo in semana_df.groupby("empleado"):
            grupo_ordenado = grupo.sort_values("fecha")
            tabla_emp = grupo_ordenado[["dia", "fecha", "hora_entrada", "hora_salida", "horas_brutas", "horas_normales"]].copy()
            tabla_emp.columns = ["Día", "Fecha", "Entrada", "Salida", "Horas brutas", "Horas normales"]

            # Nombre de hoja válido y único (máx 31 caracteres, sin caracteres especiales)
            nombre_hoja = "".join(ch for ch in nombre_emp if ch not in r'[]:*?/\\')[:31] or "Empleado"
            base = nombre_hoja
            contador = 1
            while nombre_hoja in nombres_hoja_usados:
                sufijo = f"_{contador}"
                nombre_hoja = base[:31 - len(sufijo)] + sufijo
                contador += 1
            nombres_hoja_usados.add(nombre_hoja)

            tabla_emp.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)

            ws_emp = writer.sheets[nombre_hoja]
            ws_emp["A1"] = nombre_emp
            ws_emp["A1"].font = Font(bold=True, size=13)
            ws_emp.merge_cells("A1:F1")

            fila_header = 2
            for col_idx in range(1, 7):
                celda = ws_emp.cell(row=fila_header, column=col_idx)
                celda.font = fuente_header
                celda.fill = azul_oscuro
                celda.alignment = centrado

            fila_total = fila_header + len(tabla_emp) + 1
            ws_emp.cell(row=fila_total, column=1, value="TOTAL SEMANA").font = fuente_negrita
            for col_idx, col_nombre in [(5, "Horas brutas"), (6, "Horas normales")]:
                celda_total = ws_emp.cell(row=fila_total, column=col_idx, value=f"=SUM({get_column_letter(col_idx)}{fila_header+1}:{get_column_letter(col_idx)}{fila_header+len(tabla_emp)})")
                celda_total.font = fuente_negrita
                celda_total.fill = amarillo
                celda_total.alignment = centrado

            anchos = [12, 12, 10, 10, 13, 15]
            for col_idx, ancho in enumerate(anchos, start=1):
                ws_emp.column_dimensions[get_column_letter(col_idx)].width = ancho

    st.download_button(
        "⬇️ Descargar semana en Excel (por empleado)",
        data=buffer.getvalue(),
        file_name=f"horario_semana_{fecha_inicio_semana}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.divider()
    st.markdown("**⚠️ Eliminar toda la semana**")
    st.caption(f"Borra los {len(semana_df)} registro(s) de TODOS los empleados entre el {fecha_inicio_semana.strftime('%d/%m/%Y')} y el {fecha_fin_semana.strftime('%d/%m/%Y')}. Esta acción no se puede deshacer.")
    confirmar_borrado_semana = st.checkbox("Sí, quiero eliminar todos los registros de esta semana")
    if st.button("🗑️ Eliminar semana completa", disabled=not confirmar_borrado_semana, use_container_width=True):
        with get_connection() as conn:
            execute(conn, "DELETE FROM horas WHERE fecha BETWEEN ? AND ?",
                    (str(fecha_inicio_semana), str(fecha_fin_semana)))
        st.success(f"Se eliminaron los registros de la semana del {fecha_inicio_semana.strftime('%d/%m/%Y')}.")
        st.rerun()

    st.divider()
    st.markdown("**Editar o eliminar un registro**")
    registro_id = st.selectbox(
        "Selecciona un registro",
        semana_df["id"].tolist(),
        format_func=lambda x: (
            f"{semana_df[semana_df['id']==x]['empleado'].values[0]} — "
            f"{nombre_dia(semana_df[semana_df['id']==x]['fecha'].values[0])} "
            f"{str(semana_df[semana_df['id']==x]['fecha'].values[0])[:10]}"
        )
    )
    registro = semana_df[semana_df["id"] == registro_id].iloc[0]

    with st.form("editar_registro_semana"):
        col_e, col_s = st.columns(2)
        with col_e:
            try:
                entrada_edit_default = datetime.strptime(str(registro["hora_entrada"])[:5], "%H:%M").time()
            except Exception:
                entrada_edit_default = time(8, 0)
            entrada_edit = st.time_input("Hora entrada", value=entrada_edit_default)
        with col_s:
            try:
                salida_edit_default = datetime.strptime(str(registro["hora_salida"])[:5], "%H:%M").time()
            except Exception:
                salida_edit_default = time(17, 0)
            salida_edit = st.time_input("Hora salida", value=salida_edit_default)

        col_guardar, col_eliminar = st.columns(2)
        guardar_edicion = col_guardar.form_submit_button("💾 Guardar cambios", use_container_width=True)
        eliminar_registro = col_eliminar.form_submit_button("🗑️ Eliminar este registro", use_container_width=True)

    if guardar_edicion:
        entrada_dt = datetime.combine(date.today(), entrada_edit)
        salida_dt = datetime.combine(date.today(), salida_edit)
        horas_normales_edit = min(max(0, (salida_dt - entrada_dt).total_seconds() / 3600), horas_normales_max)
        with get_connection() as conn:
            execute(conn, """
                UPDATE horas SET hora_entrada=?, hora_salida=?, horas_normales=? WHERE id=?
            """, (str(entrada_edit), str(salida_edit), round(horas_normales_edit, 4), int(registro_id)))
        st.success("Registro actualizado.")
        st.rerun()

    if eliminar_registro:
        with get_connection() as conn:
            execute(conn, "DELETE FROM horas WHERE id=?", (int(registro_id),))
        st.success("Registro eliminado.")
        st.rerun()
